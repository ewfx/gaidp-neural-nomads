import os
import json
import csv
import sqlite3
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
import argparse
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill

class SQLiteValidator:

    
    def __init__(self, 
                 db_path: str, 
                 log_file: str = "validation.log"):

        # Setup logging
        self.setup_logging(log_file)
        
        # Validate database path
        if not os.path.exists(db_path):
            self.logger.error(f"Database file not found: {db_path}")
            raise FileNotFoundError(f"SQLite database file not found: {db_path}")
        
        self.db_path = db_path
        self.logger.info(f"Initialized validator for database: {db_path}")
    
    def setup_logging(self, log_file: str):

        # Ensure log directory exists
        log_dir = os.path.dirname(log_file)
        if log_dir:
            os.makedirs(log_dir, exist_ok=True)
        
        # Configure logging
        logging.basicConfig(
            level=logging.DEBUG,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        
        self.logger = logging.getLogger(__name__)
    
    def _connect_database(self) -> sqlite3.Connection:

        try:
            conn = sqlite3.connect(self.db_path)
            return conn
        except sqlite3.Error as e:
            self.logger.error(f"Database connection error: {e}")
            raise
    
    def execute_validation_query(self, 
                             query: str, 
                             rule_id: str, 
                             rule_name: str) -> List[Dict[str, Any]]:

        try:
            # Establish database connection
            conn = self._connect_database()
            cursor = conn.cursor()
            
            # Execute the query
            self.logger.info(f"Executing rule: {rule_id} - {rule_name}")
            print(f"Executing query: {query}")
            cursor.execute(query)
            
            
            # Collect failures
            failures = []
            for row in cursor.fetchall():                
                # Check if the row contains a valid transaction ID
                if row and len(row) > 0:
                    # Try to use the first column as transaction ID
                    transaction_id = str(row[0])
                    
                    failures.append({
                        "transaction_id": transaction_id,
                        "rule_id": rule_id,
                        "rule_name": rule_name
                    })
            
            # Close connection
            cursor.close()
            conn.close()
            
            self.logger.info(f"Rule {rule_id} found {len(failures)} violations")
            return failures
        
        except sqlite3.Error as e:
            self.logger.error(f"Error executing rule {rule_id}: {e}")
            return []
        except Exception as e:
            self.logger.error(f"Unexpected error in rule {rule_id}: {e}")
            return []



    def load_validation_rules(self, rules_file: str) -> Dict[str, Any]:

        try:
            with open(rules_file, 'r', encoding='utf-8') as f:
                rules = json.load(f)
            
            # Filter active rules
            active_rules = {
                k: v for k, v in rules.items() 
                if v.get('status', '').lower() == 'active'
            }
            
            self.logger.info(f"Loaded {len(active_rules)} active validation rules")
            return active_rules
        
        except FileNotFoundError:
            self.logger.error(f"Rules file not found: {rules_file}")
            raise
        except json.JSONDecodeError:
            self.logger.error(f"Invalid JSON in rules file: {rules_file}")
            raise


    def validate_data(self, 
                  rules_file: str, 
                  output_file: str = "validation_results.csv", 
                  excel_output_file: str = "validation_results.xlsx", 
                  original_file: str = "../Temp_files/new_tran.csv",
                  identifier:str = "fed_default") -> Dict[str, Any]:

        try:
            start_time = datetime.now()
            
            # Load validation rules
            rules = self.load_validation_rules(rules_file)

            # Calculate total number of transactions dynamically
            conn = self._connect_database()
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM transactions")
            total_transactions = cursor.fetchone()[0]
            cursor.close()
            conn.close()
            
            cutoff = 0.49 * total_transactions
            # Collect all failures
            all_failures = []
            
            # Track rules with universal failures
            universal_failure_rules = []
            
            # Detailed rule performance tracking
            rule_performance = []
            
            # Execute each rule sequentially
            for rule_id, rule_data in rules.items():
                query = rule_data.get('sql_query', '')
                rule_name = rule_data.get('rule_name', rule_id)
                rule_description = rule_data.get('description', 'No description')
                
                if not query:
                    self.logger.warning(f"No SQL query for rule {rule_id}")
                    continue
                
                # Execute individual rule
                start_rule_time = datetime.now()
                rule_failures = self.execute_validation_query(
                    query, 
                    rule_id, 
                    rule_name
                )
                end_rule_time = datetime.now()
                
                # Track rule performance
                rule_performance.append({
                    "rule_id": rule_id,
                    "rule_name": rule_name,
                    "rule_description": rule_description,
                    "failures": len(rule_failures),
                    "failure_rate": round(len(rule_failures) / total_transactions * 100, 2),
                    "execution_time": (end_rule_time - start_rule_time).total_seconds()
                })
                
                # Check if the rule fails for all transactions
                if len(rule_failures) >= cutoff:
                    universal_failure_rules.append({
                        "rule_id": rule_id,
                        "rule_name": rule_name,
                        "rule_description": rule_description,
                        "sql_query": query
                    })
                    continue
                
                all_failures.extend(rule_failures)
            
            # Group failures by transaction
            transaction_failures = self._group_failures(all_failures)
            
            # Export results
            self._export_to_csv(transaction_failures, output_file)
            self._export_to_xlsx(transaction_failures, original_file, excel_output_file)
            
            end_time = datetime.now()
            
            # Return validation summary with enhanced metadata
            return {
                "timestamp": start_time.isoformat(),
                "total_rules": len(rules),
                "total_transactions": total_transactions,
                "failed_transactions": len(transaction_failures),
                "total_failures": len(all_failures),
                "failure_rate": round(len(transaction_failures) / total_transactions * 100, 2),
                "output_file": output_file,
                "excel_output_file": excel_output_file,
                "execution_time": (end_time - start_time).total_seconds(),
                "rule_performance": rule_performance,
                "universal_failure_rules": universal_failure_rules, 
                "identifier" : identifier
            }
        
        except Exception as e:
            self.logger.error(f"Validation process failed: {e}")
            return {"error": str(e)}
    
    def _group_failures(self, failures: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, str]]]:

        transaction_failures = {}
        for failure in failures:
            transaction_id = failure['transaction_id']
            rule_id = failure['rule_id']
            rule_name = failure['rule_name']
            if transaction_id not in transaction_failures:
                transaction_failures[transaction_id] = []
            
            transaction_failures[transaction_id].append({
                "rule_id": rule_id,
                "rule_name": rule_name
            })
        
        return transaction_failures

    def _export_to_csv(self, 
                       transaction_failures: Dict[str, List[str]], 
                       output_file: str):

        try:
            with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(["transaction_id", "failed_rules"])
                
                for transaction_id, failed_rules in transaction_failures.items():
                    verbose_rules = " && ".join([f"Failed rule: {rule}" for rule in failed_rules])
                    writer.writerow([transaction_id, verbose_rules])
            
            self.logger.info(f"Exported {len(transaction_failures)} transaction failures to {output_file}")
        
        except Exception as e:
            self.logger.error(f"Error exporting to CSV: {e}")
            raise

    def _export_to_xlsx(self, 
                        transaction_failures: Dict[str, List[Dict[str, str]]], 
                        original_file: str, 
                        output_file: str):

        try:
            wb = openpyxl.Workbook()
            ws = wb.active

            # Load original data
            with open(original_file, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)

            # Apply color coding and add explanations
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                transaction_id = row[0].value
                if transaction_id in transaction_failures:
                    for cell in row:
                        cell.fill = red_fill
                    verbose_rules = " && ".join([f"Failed rule: {rule['rule_name']} ({rule['rule_id']})" for rule in transaction_failures[transaction_id]])
                    row[-1].value = verbose_rules

            wb.save(output_file)
            self.logger.info(f"Exported {len(transaction_failures)} transaction failures to {output_file}")
        
        except Exception as e:
            self.logger.error(f"Error exporting to Excel: {e}")
            raise

